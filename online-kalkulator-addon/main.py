from fastapi import FastAPI
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List
from openpyxl import Workbook
from io import BytesIO

app = FastAPI()

class Item(BaseModel):
    name: str
    price: float
    qty: float

class Payload(BaseModel):
    items: List[Item]

@app.post("/generate-excel")
def generate_excel(payload: Payload):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Name", "Price", "Qty", "Total"])

    for idx, item in enumerate(payload.items, start=2):
        ws[f"A{idx}"] = item.name
        ws[f"B{idx}"] = item.price
        ws[f"C{idx}"] = item.qty
        ws[f"D{idx}"] = f"=B{idx}*C{idx}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=output.xlsx"}
    )
