import subprocess
import platform
import tempfile
import os
import zipfile
import re
from datetime import datetime
from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import range_boundaries

origins = [
    "https://online-kalkulator.at",  # Production
    "http://localhost:8000",             # Local Frontend Dev
]

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["POST", "GET"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"] 
)

if platform.system() == "Windows":
    LIBREOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"
    WORKDIR = r"C:\tmp\workdir"
    APPDIR = "./"
else:
    LIBREOFFICE = "libreoffice"
    WORKDIR = "/data/workdir"
    APPDIR = "/app"

os.makedirs(WORKDIR, exist_ok=True)

class Maschinenkosten(BaseModel):
    Aktiv: int
    Name: str
    Anschaffungspreis: float
    Restwert: float
    Zinssatz: float
    Nutzungsdauer: float
    Jahresstunden: float
    MieteAbs: float
    GebuehrenRel: float
    VersicherungRel: float
    UnterbringungRel: float
    TreibstoffVerbrauch: float
    TreibstoffKosten: float
    SchmierstoffVerbrauch: float
    SchmierstoffKosten: float
    SonstigesVerbrauch: float
    SonstigesKosten: float
    ReparaturenRel: float

class Maschinenkalkulation(BaseModel):
    Personalkosten: float
    NebenzeitenPersonal: float
    NebenzeitenGrundmaschine: float
    OrganisationsUnternehmenskosten: float
    Wagnis: float
    LeistungsbereichMin: float
    LeistungsbereichMax: float

class Arbeitsverfahren(BaseModel):
    GM: Maschinenkosten
    M1: Maschinenkosten
    M2: Maschinenkosten
    GK: Maschinenkalkulation
    version: int

    # This config block provides the default data for Swagger UI
    model_config = {
        "json_schema_extra": {
            "example": {
                "GM": {
                    "Aktiv": 1, "Name": "Rübenvollernter-Selbstfahrer", "Anschaffungspreis": 600000,
                    "Restwert": 180000, "Zinssatz": 0.04, "Nutzungsdauer": 10, "Jahresstunden": 350,
                    "MieteAbs": 0, "GebuehrenRel": 0.001, "VersicherungRel": 0.005, "UnterbringungRel": 0.005,
                    "TreibstoffVerbrauch": 70, "TreibstoffKosten": 1.4, "SchmierstoffVerbrauch": 0.2,
                    "SchmierstoffKosten": 4, "SonstigesVerbrauch": 0, "SonstigesKosten": 0, "ReparaturenRel": 1.5
                },
                "M1": {
                    "Aktiv": 0, "Name": "", "Anschaffungspreis": 0, "Restwert": 0, "Zinssatz": 0,
                    "Nutzungsdauer": 0, "Jahresstunden": 0, "MieteAbs": 0, "GebuehrenRel": 0,
                    "VersicherungRel": 0, "UnterbringungRel": 0, "TreibstoffVerbrauch": 0,
                    "TreibstoffKosten": 0, "SchmierstoffVerbrauch": 0, "SchmierstoffKosten": 0,
                    "SonstigesVerbrauch": 0, "SonstigesKosten": 0, "ReparaturenRel": 0
                },
                "M2": {
                    "Aktiv": 0, "Name": "", "Anschaffungspreis": 0, "Restwert": 0, "Zinssatz": 0,
                    "Nutzungsdauer": 0, "Jahresstunden": 0, "MieteAbs": 0, "GebuehrenRel": 0,
                    "VersicherungRel": 0, "UnterbringungRel": 0, "TreibstoffVerbrauch": 0,
                    "TreibstoffKosten": 0, "SchmierstoffVerbrauch": 0, "SchmierstoffKosten": 0,
                    "SonstigesVerbrauch": 0, "SonstigesKosten": 0, "ReparaturenRel": 0
                },
                "GK": {
                    "Personalkosten": 32, "NebenzeitenPersonal": 0.25, "NebenzeitenGrundmaschine": 0.1,
                    "OrganisationsUnternehmenskosten": 0.12, "Wagnis": 0.03,
                    "LeistungsbereichMin": 1.5, "LeistungsbereichMax": 2.5
                },
                "version": 2
            }
        }
    }

def set_named_value(ws: Worksheet, name_symbol, new_value):
    """
    Updates the value of a symbolic name in an openpyxl workbook.
    Handles both cell-based named ranges and named constants.
    """
    if name_symbol not in ws.defined_names:
        raise ValueError(f"Named symbol '{name_symbol}' not found in workbook.")

    defn = ws.defined_names[name_symbol]
    dests = list(defn.destinations)

    if dests:
        # Case 1: The name refers to a cell or range (Named Cell)
        for sheet_title, cell_address in dests:
            #ws = ws[sheet_title]
            # If it's a single cell reference like $A$1
            if ":" not in cell_address:
                ws[cell_address] = new_value
            else:
                # If it's a range like $A$1:$B$2, updates all cells in that range
                for row in ws[cell_address]:
                    for cell in row:
                        cell.value = new_value
    else:
        # Case 2: The name is a constant or formula (Named Constant)
        # Directly overwrite the definition with the new numeric value
        ws.defined_names[name_symbol] = DefinedName(name_symbol, attr_text=str(new_value))

def hide_row_by_named_variable(ws: Worksheet, variable_name: str, hide: bool = True):
    """
    Hides the row(s) associated with a specific named range within a worksheet.
    """

    defn = ws.defined_names.get(variable_name)

    if not defn:
        # Name not found, exit quietly or raise ValueError
        return

    # 3. Iterate through destinations (a name can point to multiple areas)
    for sheet_title, cell_address in defn.destinations:
        # Only hide if the destination is on the current worksheet
        if sheet_title == ws.title:
            # range_boundaries returns (min_col, min_row, max_col, max_row)
            # e.g., '$A$10' -> (1, 10, 1, 10)
            res = range_boundaries(cell_address)
            min_row, max_row = res[1], res[3]
            
            # Type safety check: ensure the indices are integers
            if isinstance(min_row, int) and isinstance(max_row, int):
                for r in range(min_row, max_row + 1):
                    ws.row_dimensions[r].hidden = hide

@app.post("/generate-excel-pdf")
def generate_excel_pdf(payload: Arbeitsverfahren,
                       filename: Optional[str] = Query(None)
                       ):
    # -----------------------------
    # 0. GENERATE TIMESTAMPED NAME
    # -----------------------------
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = filename or f"REPORT_{timestamp}"
    base_name = re.sub(r'[^\w\-_]', '_', base_name)

    template_path = os.path.join(APPDIR, f"Template.xlsx")
    report_dir = os.path.join(WORKDIR, f"REPORT_{timestamp}")
    os.makedirs(report_dir, exist_ok=True)
    excel_path = os.path.join(report_dir, f"{base_name}.xlsx")
    pdf_path   = os.path.join(report_dir, f"{base_name}.pdf")
    json_path  = os.path.join(report_dir, f"{base_name}.json")

    with open(json_path, "w", encoding="utf-8") as f: 
        f.write(payload.model_dump_json(indent=4))
    # -----------------------------
    # 1. CREATE EXCEL WORKBOOK
    # -----------------------------
    wb = load_workbook(template_path)
    ws = wb.active
    if ws is None:
        raise ValueError("The workbook contains no active worksheet.")
    
    sheet_name = ws.title
    
    # 1. Load your nested JSON data
    #with open(data_path, 'r', encoding='utf-8') as f:
    #    data = json.load(f)

    groups = ["GM", "M1", "M2", "GK"]

    for group in groups:
        model_instance = getattr(payload, group)
        for key, value in model_instance.model_dump().items():
            target: str = f"{group}_{key}"
            if key == "Aktiv":
                # Hide all inactive rows 
                if value == 0:
                    hide_row_by_named_variable(ws, target, True)
            else:
                set_named_value(ws, target, value)


    # Remove all row breaks if M1 and M2 are both inactive (there's enough space on a page for printing!)
    if payload.M1.Aktiv == 0 and payload.M2.Aktiv == 0:
        ws.row_breaks.brk = []
            
    #set_named_value(wb, "GM_Anschaffungspreis", 443322)


    #ws.append(["Name", "Price", "Qty", "Total"])
#
    #for idx, item in enumerate(payload.items, start=2):
    #    ws[f"A{idx}"] = item.name
    #    ws[f"B{idx}"] = item.price
    #    ws[f"C{idx}"] = item.qty
    #    ws[f"D{idx}"] = f"=B{idx}*C{idx}"

    wb.save(excel_path)

    # -----------------------------
    # 2. RECALCULATE FORMULAS AND REWRITE XLSX
    # -----------------------------

    # Create a temp directory inside the same folder
    with tempfile.TemporaryDirectory(dir=WORKDIR) as tmpdir:

        # Run LibreOffice conversion inside the temp dir
        subprocess.run(
            [
                LIBREOFFICE,
                "--headless",
                "--convert-to", "xlsx",
                "--outdir", tmpdir,
                excel_path
            ],
            check=True
        )

        # Find the converted file
        converted = [
            f for f in os.listdir(tmpdir)
            if f.endswith(".xlsx")
        ][0]

        converted_path = os.path.join(tmpdir, converted)

        # Replace original file with recalculated one
        os.replace(converted_path, excel_path)

        # Temp directory is automatically deleted here


    if platform.system() == "Windows":
        # The ':Zone.Identifier' suffix targets the Alternate Data Stream
        with open(f"{excel_path}:Zone.Identifier", "w") as f:
            f.write("[ZoneTransfer]\nZoneId=3")

    # -----------------------------
    # 3. CONVERT RECALCULATED XLSX → PDF
    # -----------------------------
    subprocess.run(
        [
            LIBREOFFICE,
            "--headless",
            "--calc",
            "--convert-to", "pdf",
            "--outdir", report_dir,
            excel_path
        ],
        check=True
    )

    # -----------------------------
    # 4. PACKAGE BOTH FILES INTO ZIP
    # -----------------------------

    zip_path = os.path.join(WORKDIR, f"{base_name}.zip") 
    with open(zip_path, "wb") as file_stream: 
        with zipfile.ZipFile(file_stream, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(json_path,  arcname=f"{base_name}.json")
            zipf.write(excel_path, arcname=f"{base_name}.xlsx")
            zipf.write(pdf_path,   arcname=f"{base_name}.pdf")

    if platform.system() != "Windows":
        os.remove(json_path)
        os.remove(excel_path)
        os.remove(pdf_path)
        os.rmdir(report_dir)

    # Automatically handles the file opening, streaming, 
    # and Content-Disposition header
    return FileResponse(
        path=zip_path, 
        filename=f"{base_name}.zip", 
        media_type="application/zip"
    )

    return StreamingResponse(
        open(zip_path, "rb"),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={base_name}.zip"}
    )
